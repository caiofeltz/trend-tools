import pandas as pd
import os
import sys
import argparse
from pytrends.request import TrendReq
import matplotlib.pyplot as plt
import requests
import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io
import time
import random


class TrendsBatchAnalyzer:
    """Analyzes multiple keywords from a spreadsheet and generates Excel report."""
    
    def __init__(self):
        self.output_dir = "output"
        os.makedirs(self.output_dir, exist_ok=True)
        self.errors = []
        self.results = {}
    
    def validate_input_data(self, df):
        """
        Validates input spreadsheet.
        Required column: keyword
        Optional columns: geo, timeframe, language
        Returns: Validated DataFrame or raises error
        """
        if df is None or df.empty:
            raise ValueError("Planilha vazia")
        
        # Check for required column
        if 'keyword' not in df.columns:
            raise ValueError("Coluna 'keyword' obrigatória")
        
        # Normalize column names
        df.columns = df.columns.str.lower().str.strip()
        
        # Check for NaN in keyword column
        if df['keyword'].isna().any():
            raise ValueError("Coluna 'keyword' contém valores vazios")
        
        # Set defaults for optional columns
        if 'geo' not in df.columns:
            df['geo'] = 'BR'
        else:
            df['geo'] = df['geo'].fillna('BR')
        
        if 'timeframe' not in df.columns:
            df['timeframe'] = 'today 12-m'
        else:
            df['timeframe'] = df['timeframe'].fillna('today 12-m')
        
        if 'language' not in df.columns:
            df['language'] = 'pt'
        else:
            df['language'] = df['language'].fillna('pt')
        
        # Clean whitespace
        df['keyword'] = df['keyword'].str.strip()
        df['geo'] = df['geo'].str.strip().str.upper()
        df['timeframe'] = df['timeframe'].str.strip()
        df['language'] = df['language'].str.strip().str.lower()
        
        # Validate geo codes (basic validation)
        valid_geos = ['', 'US', 'BR', 'MX', 'GB', 'DE', 'FR', 'IT', 'ES', 'JP', 'CN', 'IN', 'AU', 'CA', 'AR', 'PT']
        for geo in df['geo'].unique():
            if geo not in valid_geos:
                print(f"⚠️  Aviso: Código geo '{geo}' pode não ser válido")
        
        # Validate language codes
        valid_langs = ['pt', 'en', 'es', 'fr', 'de', 'it', 'ja', 'zh', 'ru', 'ar', 'hi']
        for lang in df['language'].unique():
            if lang not in valid_langs:
                print(f"⚠️  Aviso: Código de linguagem '{lang}' pode não ser válido")
        
        # Check for duplicate keywords
        duplicates = df[df.duplicated(subset=['keyword'], keep=False)]['keyword'].unique()
        if len(duplicates) > 0:
            print(f"\n⚠️  Aviso: {len(duplicates)} keyword(s) duplicada(s) encontrada(s):")
            for dup in duplicates:
                count = len(df[df['keyword'] == dup])
                print(f"   - '{dup}' aparece {count} vezes")
            print("   → Removendo duplicatas, mantendo primeira ocorrência\n")
            df = df.drop_duplicates(subset=['keyword'], keep='first')
        
        return df
    
    def fetch_trends_data(self, keyword, geo, timeframe):
        """
        Fetches Interest Over Time and Interest by Region from Google Trends.
        Implements retry logic with exponential backoff for rate limiting.
        """
        max_retries = 3
        base_delay = 3
        
        for attempt in range(max_retries):
            try:
                # Add randomized delay to avoid rate limiting
                if attempt == 0:
                    delay = base_delay + random.uniform(1, 3)
                else:
                    delay = base_delay * (2 ** attempt) + random.uniform(0, 2)
                
                if attempt > 0:
                    print(f"  ⏳ Tentativa {attempt + 1}/{max_retries} (aguardando {delay:.1f}s)...")
                    time.sleep(delay)
                
                print(f"  → Buscando Google Trends: {keyword} (geo={geo}, timeframe={timeframe})")
                
                # Create TrendReq instance
                pytrends = TrendReq(hl='en-US', tz=360)
                kw_list = [keyword]
                
                pytrends.build_payload(kw_list, cat=0, timeframe=timeframe, geo=geo, gprop='')
                
                # Interest Over Time
                interest_over_time_df = pytrends.interest_over_time()
                if not interest_over_time_df.empty:
                    interest_over_time_df = interest_over_time_df.drop(labels=['isPartial'], axis='columns')
                
                # Delay between requests to avoid rate limiting
                time.sleep(random.uniform(1, 2))
                
                # Interest By Region
                interest_by_region_df = pytrends.interest_by_region(resolution='REGION', inc_low_vol=True, inc_geo_code=False)
                
                print(f"  ✓ Google Trends obtido com sucesso")
                return interest_over_time_df, interest_by_region_df
                
            except Exception as e:
                error_msg = str(e)
                
                # Check if it's a rate limit error
                if '429' in error_msg or 'Too Many Requests' in error_msg or 'quota' in error_msg.lower():
                    if attempt < max_retries - 1:
                        print(f"  ⚠️  Rate limit (429) detectado! Aguardando antes de nova tentativa...")
                        continue
                    else:
                        self.errors.append(f"Google Trends Error ({keyword}): Rate limit excedido após {max_retries} tentativas")
                        print(f"  ✗ Erro: Rate limit não resolvido após {max_retries} tentativas")
                        return pd.DataFrame(), pd.DataFrame()
                else:
                    # For other errors, retry once
                    if attempt == 0 and 'Connection' not in error_msg:
                        print(f"  ⚠️  Erro temporário: {error_msg[:50]}... Tentando novamente...")
                        time.sleep(2)
                        continue
                    
                    self.errors.append(f"Google Trends Error ({keyword}): {error_msg}")
                    print(f"  ✗ Erro Google Trends: {error_msg[:60]}")
                    return pd.DataFrame(), pd.DataFrame()
    
    def fetch_wikipedia_data(self, keyword, lang='pt', days=365, start_date_str=None, end_date_str=None):
        """
        Fetches daily pageviews for a Wikipedia article.
        Implements retry logic for robustness.
        """
        max_retries = 2
        
        for attempt in range(max_retries):
            try:
                print(f"  → Buscando Wikipedia: {keyword} ({lang})")
                
                if start_date_str and end_date_str:
                    start_str = start_date_str.replace('-', '')
                    end_str = end_date_str.replace('-', '')
                else:
                    end_date = datetime.datetime.now()
                    start_date = end_date - datetime.timedelta(days=days)
                    start_str = start_date.strftime("%Y%m%d")
                    end_str = end_date.strftime("%Y%m%d")
                
                article_title = keyword.replace(' ', '_')
                url = f"https://wikimedia.org/api/rest_v1/metrics/pageviews/per-article/{lang}.wikipedia/all-access/all-agents/{article_title}/daily/{start_str}/{end_str}"
                
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                }
                
                response = requests.get(url, headers=headers, timeout=15)
                
                if response.status_code == 200:
                    data = response.json()
                    items = data.get('items', [])
                    
                    if not items:
                        print(f"  ⚠️  Nenhum dado Wikipedia encontrado")
                        return pd.DataFrame()
                    
                    dates = [datetime.datetime.strptime(item['timestamp'], "%Y%m%d00") for item in items]
                    views = [item['views'] for item in items]
                    
                    df = pd.DataFrame({'views': views}, index=dates)
                    print(f"  ✓ Wikipedia obtido com sucesso")
                    return df
                elif response.status_code == 404:
                    print(f"  ⚠️  Página Wikipedia não encontrada")
                    return pd.DataFrame()
                elif response.status_code == 429:
                    if attempt < max_retries - 1:
                        print(f"  ⚠️  Rate limit Wikipedia detectado. Aguardando...")
                        time.sleep(random.uniform(3, 5))
                        continue
                    else:
                        print(f"  ⚠️  Wikipedia rate limit (aguarde e tente novamente)")
                        return pd.DataFrame()
                else:
                    print(f"  ⚠️  Wikipedia API Error: {response.status_code}")
                    return pd.DataFrame()
                    
            except requests.exceptions.Timeout:
                if attempt < max_retries - 1:
                    print(f"  ⚠️  Timeout Wikipedia. Tentando novamente...")
                    time.sleep(2)
                    continue
                else:
                    print(f"  ⚠️  Wikipedia timeout")
                    return pd.DataFrame()
            except Exception as e:
                print(f"  ⚠️  Wikipedia Error: {str(e)[:60]}")
                return pd.DataFrame()
        
        return pd.DataFrame()
    
    def parse_timeframe_for_dates(self, timeframe):
        """Parses timeframe string to extract start and end dates."""
        try:
            parts = timeframe.split()
            if len(parts) == 2:
                datetime.datetime.strptime(parts[0], "%Y-%m-%d")
                datetime.datetime.strptime(parts[1], "%Y-%m-%d")
                return parts[0], parts[1]
        except (ValueError, IndexError):
            pass
        return None, None
    
    def plot_trends(self, google_df, wiki_df, keyword):
        """
        Plots Interest Over Time (Google) and Pageviews (Wikipedia).
        Returns image as bytes or None.
        """
        try:
            if (google_df is None or google_df.empty) and (wiki_df is None or wiki_df.empty):
                return None
            
            fig, ax1 = plt.subplots(figsize=(12, 6))
            
            # Plot Google Trends on primary y-axis
            if google_df is not None and not google_df.empty:
                ax1.plot(google_df.index, google_df[keyword], color='tab:blue', label='Google Search Interest', linewidth=2)
                ax1.set_xlabel('Date', fontsize=10)
                ax1.set_ylabel('Google Interest (0-100)', color='tab:blue', fontsize=10)
                ax1.tick_params(axis='y', labelcolor='tab:blue')
            
            # Plot Wikipedia Views on secondary y-axis
            if wiki_df is not None and not wiki_df.empty:
                ax2 = ax1.twinx()
                ax2.plot(wiki_df.index, wiki_df['views'], color='tab:orange', linestyle='--', label='Wikipedia Pageviews', linewidth=2)
                ax2.set_ylabel('Wikipedia Pageviews', color='tab:orange', fontsize=10)
                ax2.tick_params(axis='y', labelcolor='tab:orange')
            
            plt.title(f'Interest & Pageviews Over Time: {keyword}', fontsize=12, fontweight='bold')
            fig.tight_layout()
            plt.grid(True, axis='x', alpha=0.3)
            
            # Manual Legend
            lines1, labels1 = ax1.get_legend_handles_labels() if google_df is not None and not google_df.empty else ([], [])
            if wiki_df is not None and not wiki_df.empty:
                lines2, labels2 = ax2.get_legend_handles_labels()
            else:
                lines2, labels2 = [], []
            
            ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left')
            
            # Save to bytes
            img_bytes = io.BytesIO()
            plt.savefig(img_bytes, format='png', dpi=100, bbox_inches='tight')
            img_bytes.seek(0)
            plt.close()
            
            return img_bytes
        except Exception as e:
            print(f"  ✗ Erro ao gerar gráfico: {e}")
            return None
    
    def process_keyword(self, row, keyword_index=0, total_keywords=1):
        """
        Processes a single keyword row.
        Returns: Dictionary with results and image
        """
        keyword = row['keyword']
        geo = row['geo']
        timeframe = row['timeframe']
        language = row['language']
        
        # Calculate adaptive delay based on position in batch
        if keyword_index > 0:
            # Increase delay for keywords later in the batch to avoid rate limiting
            base_delay = 2 + (keyword_index * 0.5)
            total_delay = base_delay + random.uniform(1, 3)
            print(f"\n⏳ Aguardando {total_delay:.1f}s para evitar rate limit...")
            time.sleep(total_delay)
        
        print(f"\n📊 Processando ({keyword_index + 1}/{total_keywords}): {keyword}")
        
        result = {
            'keyword': keyword,
            'geo': geo,
            'timeframe': timeframe,
            'language': language,
            'time_df': pd.DataFrame(),
            'region_df': pd.DataFrame(),
            'wiki_df': pd.DataFrame(),
            'image': None,
            'success': False
        }
        
        try:
            # Fetch Google Trends
            time_df, region_df = self.fetch_trends_data(keyword, geo, timeframe)
            result['time_df'] = time_df
            result['region_df'] = region_df
            
            # Fetch Wikipedia Data
            wiki_start, wiki_end = self.parse_timeframe_for_dates(timeframe)
            wiki_keyword = keyword.capitalize() if keyword.islower() else keyword
            wiki_df = self.fetch_wikipedia_data(wiki_keyword, lang=language, start_date_str=wiki_start, end_date_str=wiki_end)
            result['wiki_df'] = wiki_df
            
            # Generate plot
            image = self.plot_trends(time_df, wiki_df, keyword)
            result['image'] = image
            
            result['success'] = True
            print(f"✓ Keyword processada com sucesso")
            
        except Exception as e:
            self.errors.append(f"Erro processando {keyword}: {e}")
            print(f"✗ Erro: {e}")
        
        return result
    
    def process_batch(self, input_file):
        """
        Processes all keywords from input spreadsheet.
        Returns: List of results
        """
        # Read input file
        print(f"📂 Lendo arquivo: {input_file}")
        
        try:
            if input_file.endswith('.xlsx'):
                df = pd.read_excel(input_file)
            elif input_file.endswith('.csv'):
                df = pd.read_csv(input_file)
            else:
                raise ValueError("Arquivo deve ser .xlsx ou .csv")
        except Exception as e:
            raise ValueError(f"Erro ao ler arquivo: {e}")
        
        # Validate input
        print("✓ Validando dados...")
        df = self.validate_input_data(df)
        print(f"✓ {len(df)} keywords para processar")
        
        # Process each keyword
        results = []
        total_keywords = len(df)
        for idx, row in df.iterrows():
            result = self.process_keyword(row, keyword_index=idx, total_keywords=total_keywords)
            results.append(result)
        
        return results
    
    def create_excel_report(self, results, output_file):
        """
        Creates Excel report with one sheet per keyword.
        Includes images in each sheet.
        """
        print(f"\n📄 Criando relatório Excel...")
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        for idx, result in enumerate(results):
            if not result['success']:
                continue
            
            keyword = result['keyword']
            sheet_name = keyword[:31]  # Excel limit
            ws = wb.create_sheet(title=sheet_name)
            
            # Header with parameters
            row = 1
            ws[f'A{row}'] = f"Relatório: {keyword}"
            ws[f'A{row}'].font = Font(size=14, bold=True)
            ws.merge_cells(f'A{row}:D{row}')
            
            row += 1
            ws[f'A{row}'] = f"Data: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}"
            
            row += 1
            ws[f'A{row}'] = "Parâmetros:"
            ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
            ws[f'A{row}'] = "Keyword:"
            ws[f'B{row}'] = result['keyword']
            
            row += 1
            ws[f'A{row}'] = "Região:"
            ws[f'B{row}'] = result['geo']
            
            row += 1
            ws[f'A{row}'] = "Período:"
            ws[f'B{row}'] = result['timeframe']
            
            row += 1
            ws[f'A{row}'] = "Linguagem:"
            ws[f'B{row}'] = result['language']
            
            # Insert image if available
            row += 2
            if result['image']:
                ws[f'A{row}'] = "Gráfico:"
                ws[f'A{row}'].font = Font(bold=True, size=11)
                
                row += 1
                img_path = os.path.join(self.output_dir, f"temp_plot_{idx}.png")
                with open(img_path, 'wb') as f:
                    f.write(result['image'].getvalue())
                
                xl_image = XLImage(img_path)
                xl_image.width = 600
                xl_image.height = 400
                ws.add_image(xl_image, f'A{row}')
                row += 21  # Make room for image
            
            # Google Trends Stats
            row += 2
            ws[f'A{row}'] = "Google Trends - Interesse ao Longo do Tempo"
            ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws.merge_cells(f'A{row}:D{row}')
            
            if not result['time_df'].empty:
                row += 1
                time_df = result['time_df'].head(10)
                
                # Headers
                for col_idx, col_name in enumerate(['Data'] + list(time_df.columns), 1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = col_name
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                
                # Data
                row += 1
                for idx_row, (date, values) in enumerate(time_df.iterrows()):
                    ws.cell(row=row, column=1).value = date.strftime('%d/%m/%Y') if hasattr(date, 'strftime') else str(date)
                    for col_idx, val in enumerate(values, 2):
                        ws.cell(row=row, column=col_idx).value = val
                    row += 1
            else:
                row += 1
                ws[f'A{row}'] = "Nenhum dado disponível"
            
            # Interest by Region
            row += 2
            ws[f'A{row}'] = "Interesse por Região"
            ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            ws.merge_cells(f'A{row}:D{row}')
            
            if not result['region_df'].empty:
                row += 1
                region_df = result['region_df'].head(10).sort_values(by=result['keyword'], ascending=False)
                
                # Headers
                for col_idx, col_name in enumerate(['Região'] + list(region_df.columns), 1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = col_name
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                
                # Data
                row += 1
                for idx_row, (region, values) in enumerate(region_df.iterrows()):
                    ws.cell(row=row, column=1).value = region
                    for col_idx, val in enumerate(values, 2):
                        ws.cell(row=row, column=col_idx).value = val
                    row += 1
            else:
                row += 1
                ws[f'A{row}'] = "Nenhum dado disponível"
            
            # Wikipedia Stats
            row += 2
            ws[f'A{row}'] = "Wikipedia - Visualizações"
            ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            ws.merge_cells(f'A{row}:D{row}')
            
            if not result['wiki_df'].empty:
                row += 1
                wiki_df = result['wiki_df'].tail(10)
                
                # Headers
                for col_idx, col_name in enumerate(['Data', 'Visualizações'], 1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = col_name
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                
                # Data
                row += 1
                for date, views in zip(wiki_df.index, wiki_df['views']):
                    ws.cell(row=row, column=1).value = date.strftime('%d/%m/%Y') if hasattr(date, 'strftime') else str(date)
                    ws.cell(row=row, column=2).value = views
                    row += 1
            else:
                row += 1
                ws[f'A{row}'] = "Nenhum dado disponível"
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
        
        # Save workbook
        wb.save(output_file)
        print(f"✓ Relatório salvo: {output_file}")
        
        # Clean up temporary image files
        import glob
        for temp_file in glob.glob(os.path.join(self.output_dir, "temp_plot_*.png")):
            try:
                os.remove(temp_file)
            except:
                pass
    
    def generate_report(self, input_file, output_file=None):
        """Main method to generate report from input file."""
        if output_file is None:
            output_file = os.path.join(self.output_dir, "trends_report.xlsx")
        
        try:
            # Process batch
            results = self.process_batch(input_file)
            
            # Create Excel report
            self.create_excel_report(results, output_file)
            
            # Summary
            print("\n" + "="*60)
            print("✓ RELATÓRIO CONCLUÍDO")
            print("="*60)
            successful = sum(1 for r in results if r['success'])
            print(f"Keywords processadas: {successful}/{len(results)}")
            
            if self.errors:
                print(f"\n⚠️  {len(self.errors)} erros encontrados:")
                for error in self.errors[:5]:
                    print(f"  - {error}")
                if len(self.errors) > 5:
                    print(f"  ... e mais {len(self.errors) - 5}")
            
            print(f"\n📊 Arquivo de saída: {output_file}")
            
        except Exception as e:
            print(f"\n✗ Erro crítico: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)
    
    def get_results_dataframe(self, results):
        """
        Converte resultados em DataFrame consolidado para visualização.
        Útil para Streamlit e análises.
        Retorna: DataFrame com estatísticas agregadas por keyword
        """
        summary_data = []
        
        for result in results:
            if not result['success']:
                continue
            
            keyword = result['keyword']
            
            # Extrai DataFrames
            time_df = result['time_df']
            region_df = result['region_df']
            wiki_df = result['wiki_df']
            
            # Google Trends stats - acessa primeira coluna dinamicamente
            if not time_df.empty and len(time_df.columns) > 0:
                col_name = time_df.columns[0]
                google_max = float(time_df[col_name].max())
                google_mean = float(time_df[col_name].mean())
                google_min = float(time_df[col_name].min())
            else:
                google_max = 0.0
                google_mean = 0.0
                google_min = 0.0
            
            # Region stats
            if not region_df.empty and len(region_df.columns) > 0:
                col_name = region_df.columns[0]
                top_region = region_df[col_name].idxmax()
                top_region_value = float(region_df[col_name].max())
            else:
                top_region = "N/A"
                top_region_value = 0.0
            
            # Wiki stats
            if not wiki_df.empty and 'views' in wiki_df.columns:
                wiki_max = float(wiki_df['views'].max())
                wiki_mean = float(wiki_df['views'].mean())
                wiki_min = float(wiki_df['views'].min())
            else:
                wiki_max = 0.0
                wiki_mean = 0.0
                wiki_min = 0.0
            
            summary_data.append({
                'keyword': keyword,
                'geo': result['geo'],
                'timeframe': result['timeframe'],
                'language': result['language'],
                'google_max': google_max,
                'google_mean': round(google_mean, 2),
                'google_min': google_min,
                'top_region': top_region,
                'top_region_value': top_region_value,
                'wiki_max': wiki_max,
                'wiki_mean': round(wiki_mean, 2),
                'wiki_min': wiki_min,
                'has_data': not time_df.empty
            })
        
        return pd.DataFrame(summary_data) if summary_data else pd.DataFrame()


def main():
    parser = argparse.ArgumentParser(
        description='Trends Tool - Análise em Batch de Keywords',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos de uso:
  python execution/trends_tool_batch.py input.xlsx
  python execution/trends_tool_batch.py keywords.csv -o output_report.xlsx
        """
    )
    
    parser.add_argument('input_file', type=str, help='Arquivo de entrada (.xlsx ou .csv) com keywords')
    parser.add_argument('-o', '--output', type=str, default=None, help='Arquivo de saída (padrão: output/trends_report.xlsx)')
    
    args = parser.parse_args()
    
    # Check if input file exists
    if not os.path.exists(args.input_file):
        print(f"✗ Erro: Arquivo não encontrado: {args.input_file}")
        sys.exit(1)
    
    # Run analyzer
    analyzer = TrendsBatchAnalyzer()
    analyzer.generate_report(args.input_file, args.output)


if __name__ == '__main__':
    main()
